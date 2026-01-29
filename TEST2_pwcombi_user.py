import os
import time
from openpyxl import load_workbook


device = "SWITCH"


# Jenkins 환경변수 (로컬 실행 시 fallback)
WORKSPACE = os.getenv("WORKSPACE", r"C:\Git\script_security_project")
BUILD_NO  = os.getenv("BUILD_NUMBER", "local")

result_dir = os.path.join(WORKSPACE, "RESULT")
os.makedirs(result_dir, exist_ok=True)

# 템플릿(커밋된 파일)
template_path = os.path.join(result_dir, "securitytest_template.xlsx")
if not os.path.exists(template_path):
    raise FileNotFoundError(f"Template not found: {template_path}")

# 산출물(빌드별 결과 파일)
save_path = os.path.join(result_dir, f"securitytest_result_{BUILD_NO}.xlsx")

wb = load_workbook(template_path)
sheet1 = wb["RESULT"]
sheet2 = wb["LOG"]

def get_next_row(sheet):
    for row in range(sheet.max_row, 0, -1):
        if any(cell.value for cell in sheet[row]):
            return row +1
    return 1


###전체 line 읽기###
all_lines = []

def read_all():

    global all_lines
    all_lines = []

    num_rows = 300
    num_cols = 300

    for row in range(1, num_rows + 1):
        line = crt.Screen.Get(row, 1, row, num_cols).rstrip()
        if line.strip() != "":
            all_lines.append(line)


###시험부분 line만 추출###
def select(all_lines, TEST_start, TEST_end):

    final_line = []
    capture = False

    for line in all_lines:
        if TEST_start in line:
            capture = True
            final_line.append(line)
            continue

        if capture:
            final_line.append(line)
            if TEST_end in line:
                break

    final = "\n".join(final_line)

    return final


###판단 부분###
def pw_verify(judge_count, find_line):

    read_all()
    final = select(all_lines, f"{device}# {test_name}_start", f"{device}# {test_name}_end")
    j = 0

    for judge in final.splitlines():
        if find_line in judge:
            j += 1

    result = "PASS" if j == judge_count else "FAIL"

    row1 = get_next_row(sheet1)
    row2 = get_next_row(sheet2)

    sheet1.cell(row=row1, column=2, value=test_name)
    sheet1.cell(row=row1, column=3, value=result)
    sheet1.cell(row=row1, column=4, value = j)

    sheet2.cell(row=row2, column=2, value=test_name)
    sheet2.cell(row=row2, column=3, value=final)
    

##########################################################################
###TEST1 : 1.2.1 비밀번호 보안성 기준###


judge_count = 0
combi_pw_list1 = [
        "024681357A",
        "ADGJLQETUa",
        "adgjlqetu!",
        "024681357!"
                ]

combi_pw_list2 = [
        "ADGJLQETUla!",
        "024681357Aa",
        "024681357A!",
        "024681357a!"
                ]

repeated_pw_list = [
        "AAAKTC12!",
        "aaaktc12!",
        "KTC111!",
        "KTCktc!!!"
        ]

consecutive_pw_list1 = [
        "ASDFktc1!",
        "asdfKTC1!",
        "KTC1234!#"
    ]

consecutive_pw_list2 = [
        "LKJHktc1!",
        "lkjhKTC1!",
        "KTC4321!#"
    ]

username = "admin2"
priv = "admin"

same_user = [username]




def start():
    crt.Session.Connect("/SERIAL COM5 /BAUD 115200")
    time.sleep(2)
    crt.Screen.Send("\n")
    crt.Screen.WaitForString("login")
    crt.Screen.Send("admin\n")
    crt.Screen.WaitForString("Password")
    crt.Screen.Send("Changeme1357!\r")
    crt.Screen.WaitForString("SWITCH>")
    crt.Screen.Send("\n") 
    crt.Screen.Send("enable\n") 


    
###config mode 진입###
def config_mode():

    crt.Screen.Send("conf t\n")
    crt.Screen.WaitForString(f"{device}(config)# ")


#username test 만들기
def make_username():
    crt.Screen.Synchronous = True

    #crt.Screen.Send("username test privilege 3 password")
    crt.Screen.Send(f"username {username} password {priv}")
    time.sleep(1)
    crt.Screen.Send("\r")
    crt.Screen.WaitForString("Password")
    ### HS5412 admin 계정 사용
    #crt.Screen.WaitForString("Enter admin account password")
    time.sleep(0.5)

    crt.Screen.Synchronous = False


def TEST1_PW_user(pw_case):
    crt.Screen.Send(f"{test_name}_start\n")

    time.sleep(1)
    config_mode()

    for pw in pw_case:
        make_username()
        crt.Screen.Send(f"{pw}\r")
        time.sleep(0.3)
        crt.Screen.Send("\n")
        crt.Screen.WaitForString(f"{device}(config)# ")

    # crt.Screen.Send("\n")
    time.sleep(1)
    crt.Screen.Send("end\n")
    time.sleep(0.5)
    #crt.Screen.Send("show running-config |include test\n")
    crt.Screen.Send(f"show running-config |include {username}\n")
    time.sleep(1)
    crt.Screen.Send("\n")
    #crt.Screen.Send("show username\n")
    crt.Screen.Send(f"{test_name}_end\n")
    time.sleep(1)



try:
    start()

    test_name = "TEST1_PW_combi1_user"
    TEST1_PW_user(combi_pw_list1)
    pw_verify(4, "% Your password must contain a minimum of 9 characters")

    '''
    test_name = "TEST1_PW_combi2_user"
    TEST1_PW_user(combi_pw_list2) 
    pw_verify(4, "% Your password must contain a minimum of 9 characters")

    test_name = "TEST1_PW_repeated_user"
    TEST1_PW_user(repeated_pw_list)
    pw_verify(4, "% Passwords should not have the same characters or numbers in succession.")

    test_name = "TEST1_PW_consecutive_user1"
    TEST1_PW_user(consecutive_pw_list1)
    pw_verify(3, "% Passwords should not have the consecutive characters or numbers in succession.")

    test_name = "TEST1_PW_consecutive_user2"
    TEST1_PW_user(consecutive_pw_list2)
    pw_verify(3, "% Passwords should not have the consecutive characters or numbers in succession.")

    test_name = "TEST1_PW_same_user"
    TEST1_PW_user(same_user)
    pw_verify(1, "% Passwords should not have contain a user name.")
    '''

finally:
    wb.save(save_path)