import time
from openpyxl import load_workbook

device = "HL7301"

file_path = r"C:\Users\ehh74_0i1\OneDrive\바탕 화면\2025 업무\_2025.04.03_보안기능시험_자동화\securitytest_result.xlsx"
save_path = file_path

wb = load_workbook(file_path)
sheet1 = wb["RESULT"]
sheet2 = wb["LOG"]

def get_next_row(sheet):
    for row in range(sheet.max_row, 0, -1):
        if any(cell.value for cell in sheet[row]):
            return row +1
    return 1



def login_fail(protocol, username, wrong_password):
    if protocol == "Console" or "Telnet":
        for i in range(2):
            time.sleep(1)
            crt.Screen.Send(f"{username}\r")
            time.sleep(1)
            crt.Screen.Send(f"{wrong_password}\r")
            crt.Screen.WaitForString("Username")

    elif protocol == "SSH":
        for i in range(2):
            time.sleep(1)
            crt.Screen.Send(f"{wrong_password}\r")
            crt.Screen.WaitForString("Password")

def login_success(protocol, username, right_password):
    if protocol in ["Console", "Telnet"]:
        crt.Screen.Send(f"{username}\r")
        time.sleep(1)

    crt.Screen.Send(f"{right_password}\r")
    crt.Screen.WaitForString(f"{device}>")


def start(protocol, ip, username, password):
    if protocol == "Console":
        crt.Session.Connect("/SERIAL COM6 /BAUD 115200")
        time.sleep(2)
        crt.Screen.Send("\n")
        crt.Screen.WaitForString("Username")

    elif protocol == "Telnet":
        crt.Session.Connect(f"/TELNET {ip}")

    if protocol == "SSH":
        crt.Session.Connect(f"/SSH2 {username}@{ip} /PASSWORD {password}")




def read_all():
    global all_lines
    all_lines = []

    row = 1
    while True:
        try:
            line = crt.Screen.Get(row, 1, row, 300).rstrip()
        except Exception:
            break  # 더 이상 읽을 수 없으면 종료

        if line.strip() != "":
            all_lines.append(line)

        row += 1

        # 안전장치: 너무 큰 수에서 무한루프 방지
        if row > 1000:
            break

'''
    num_rows = 300
    num_cols = 300


    for row in range(1, num_rows + 1):
        line = crt.Screen.Get(row, 1, row, num_cols).rstrip()
        if line.strip() != "":
            all_lines.append(line)
'''
    

def judge(judge_line):
    error_line = None

    # all_lines 리스트에서 Password: 찾기
    for i, line in enumerate(all_lines):
        if "Password" in line:
            # 다음 줄이 있으면 그걸 error_line으로
            if i+1 < len(all_lines):
                error_line = all_lines[i+1].strip()
            break

    result = "PASS" if error_line == judge_line else "FAIL"

    row1 = get_next_row(sheet1)
    row2 = get_next_row(sheet2)

    sheet1.cell(row=row1, column=2, value=test_name)
    sheet1.cell(row=row1, column=3, value=result)
    sheet1.cell(row=row1, column=4, value=judge_line)

    sheet2.cell(row=row2, column=2, value=test_name)
    sheet2.cell(row=row2, column=3, value="\n".join(all_lines))
    
    wb.save(save_path)


    

#test_name = "TEST5_noAuthFailureReason_Console"
#start("Console", "192.168.73.2", "admin", "Changeme1357!!")
#login_fail("Console", "admin", "Changeme1357!!")
#read_all()
#time.sleep(3)
#judge("Username: admin")
#time.sleep(3)
#login_success("Console", "admin", "Changeme1357!")

#test_name = "TEST5_noAuthFailureReason_Telnet"
#start("Telnet", "192.168.73.2", "admin", "Changeme1357!!")
#login_fail("Telnet", "admin", "Changeme1357!!")
#read_all()
#time.sleep(3)
#judge("Username: admin")
#time.sleep(3)
#login_success("Telnet", "admin", "Changeme1357!")

#test_name = "TEST5_noAuthFailureReason_SSH"
#start("SSH", "192.168.73.2", "admin", "Changeme1357!!")
#login_fail("SSH", "admin", "Changeme1357!!")
#read_all()
#time.sleep(3)
#judge("Password:")
#time.sleep(3)
#login_success("SSH", "admin", "Changeme1357!")







