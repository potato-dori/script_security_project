
import time
import re
import mss
import os
from openpyxl import load_workbook

device = "SWITCH"
file_path = r"C:\Git\script_security_project\RESULT\securitytest_result.xlsx"
save_path = file_path

wb = load_workbook(file_path)
sheet1 = wb["RESULT"]
sheet2 = wb["LOG"]

def get_next_row(sheet):
    for row in range(sheet.max_row, 0, -1):
        if any(cell.value for cell in sheet[row]):
            return row +1
    return 1


###전체 line 읽기###
all_lines = []

def read_all():
    crt.Screen.Synchronous = True
    global all_lines
    all_lines = []

    num_rows = 100
    num_cols = 300

    for row in range(1, num_rows + 1):
        line = crt.Screen.Get(row, 1, row, num_cols).rstrip()
        if line.strip() != "":
            all_lines.append(line)

    crt.Screen.Synchronous = False



###시험부분 line만 추출###
def select(all_lines, TEST_start, TEST_end):
    crt.Screen.Synchronous = True
    final_line = []
    capture = False

    for line in all_lines:
        if TEST_start in line:
            capture = True
            final_line.append(line)
            continue

        if capture:
            final_line.append(line)
            if TEST_end in line:
                break

    final = "\n".join(final_line)

    crt.Screen.Synchronous = False
    return final



###판단 부분###
def pw_verify(TEST_name, judge_count, find_line):
    crt.Screen.Synchronous = True

    read_all()
    final = '\n'.join(all_lines)
    #final = select(all_lines, f"{device}# {TEST_name}_start", f"{device}# {TEST_name}_end")
    j = 0

    for judge in final.splitlines():
        if find_line in judge:
            j += 1

    result = "PASS" if j == judge_count else "FAIL"

    row1 = get_next_row(sheet1)
    row2 = get_next_row(sheet2)

    sheet1.cell(row=row1, column=2, value=TEST_name)
    sheet1.cell(row=row1, column=3, value=result)
    sheet1.cell(row=row1, column=4, value = j)

    sheet2.cell(row=row2, column=2, value=TEST_name)
    sheet2.cell(row=row2, column=3, value=final)


    #if j == judge_count:
    #    sheet1.append([TEST_name, "PASS", j])
    #    sheet2.append([final])
    
    #else :
    #    sheet1.append([TEST_name, "FAIL", j])
    #    sheet2.append([final])
    
    wb.save(save_path)
    
    crt.Screen.Synchronous = False



###config mode 진입###
def config_mode():
    crt.Screen.Synchronous = True

    crt.Screen.Send("conf t\n")
    crt.Screen.WaitForString(f"{device}(config)# ")

    crt.Screen.Synchronous = False



##########################################################################

###TEST1 : 1.2.1 비밀번호 보안성 기준###
judge_count = 0
pw_list_combi = [
        "024681357A",
        "ADGJLQETUa",
        "adgjlqetu!",
        "024681357!",
        "ADGJLQETUla!",
        "024681357Aa",
        "024681357A!",
        "024681357a!"
                ]

repeated_pw_list = [
        "AAAKTC12!",
        "aaaktc12!",
        "KTC111!",
        "KTCktc!!!"
        ]

consecutive_pw_list = [
        "ASDFktc1!",
        "asdfKTC1!",
        "KTC1234!#",
        "LKJHktc1!",
        "lkjhKTC1!",
        "KTC4321!#"
    ]

def make_admin():
    crt.Screen.Synchronous = True

    crt.Screen.Send("\n")
    crt.Screen.WaitForString("Enter admin account password : ")

    crt.Screen.Synchronous = False


def TEST1_PW_combi(TEST1_PW_combi, test_pw):
    crt.Screen.Synchronous = True

    #crt.Screen.Send(f"{TEST1_PW_combi}_start\n")

    make_admin()

    for pw in test_pw:
        time.sleep(2)
        crt.Screen.Send(f"{pw}\r")
        crt.Screen.WaitForString("Enter admin account password : ")

    time.sleep(1)
    #crt.Screen.Send(f"{TEST1_PW_combi}_end\n")

    crt.Screen.Synchronous = False

time.sleep(1)
crt.Screen.Send("\r")

TEST1_PW_combi("TEST1_PW_combi_init", pw_list_combi)
pw_verify("TEST1_PW_combi_init", 9, "% Your password must contain a minimum of 9 characters included with at least") 
time.sleep(1)
crt.Screen.Send("\r")

TEST1_PW_combi("TEST1_PW_repeated_init", repeated_pw_list)
pw_verify("TEST1_PW_repeated_init", 4, "% Passwords should not have the same characters or numbers in succession.")
time.sleep(1)
crt.Screen.Send("\r")

TEST1_PW_combi("TEST1_PW_consecutive_init", consecutive_pw_list)
pw_verify("TEST1_PW_consecutive_init", 6, "% Passwords should not have the consecutive characters or numbers in succession.")
time.sleep(1)
crt.Screen.Send("\r")



