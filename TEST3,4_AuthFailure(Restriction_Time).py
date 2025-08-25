import time
import paramiko
from xtelnet import Telnet_Session
from openpyxl import load_workbook
from datetime import datetime

username_admin = "admin"
password_admin = "Changeme1357!"

username_admin2 = "admin2"
password_admin2 = "Changeme1357!\r"

######################################
# 결과 excel에 기록 #

device = "HL7301"
file_path = r"C:\Users\ehh74_0i1\OneDrive\바탕 화면\2025 업무\_2025.04.03_보안기능시험_자동화\securitytest_result.xlsx"
save_path = file_path

wb = load_workbook(file_path)
sheet1 = wb["RESULT"]
sheet2 = wb["LOG"]

def get_next_row(sheet):
    for row in range(sheet.max_row, 0, -1):
        if any(cell.value for cell in sheet[row]):
            return row +1
    return 1


#############################################################
# 선행 : telnet, ssh enable, username admin2, user1 생성 #

def system_enable(ip):

    telnet = Telnet_Session()
    telnet.enable_debug
    telnet.connect(ip, username=username_admin, password=f"{password_admin}\r", timeout=10)
    telnet.execute("enable\n")
    telnet.execute("con t\n")
    #telnet.execute("system enable-ssh\n")
    #telnet.execute("system enable-telnet\n")
    telnet.execute("feature telnet\n")
    telnet.execute("feature ssh\n")
    telnet.execute("end\n")
    telnet.execute("exit\n")


def make_user(ip, username, privilege):

    telnet = Telnet_Session()
    telnet.connect(ip, username=username_admin, password=f"{password_admin}\r", timeout=10)

    telnet.execute("enable\n")
    telnet.execute("con t\n")
    #telnet.execute(f"username {username} password {privilege}\r")
    telnet.execute(f"username {username} privilege {privilege} password\r")
    time.sleep(1)
    telnet.execute("Changeme1357!\r")
    time.sleep(0.5)
    telnet.execute("Changeme1357!\r")
    time.sleep(0.5)
    telnet.execute("end\n")
    telnet.execute("exit\n")

##################################################
# 차단 시간 지정 #
def set_login_time(ip, login_time):

    telnet = Telnet_Session()
    telnet.connect(ip, username=username_admin, password=f"{password_admin}\r", timeout=10)

    telnet.execute("enable\n")
    telnet.execute("con t\n")

    telnet.execute(f"user login lockout-time {login_time}\r")
    time.sleep(1)
    telnet.execute("end")
    telnet.close

##################################################
# 차단 횟수 지정 #
def set_login_times(ip, login_times):

    telnet = Telnet_Session()
    telnet.connect(ip, username=username_admin, password=f"{password_admin}\r", timeout=10)

    telnet.execute("enable\n")
    telnet.execute("con t\n")

    telnet.execute(f"user login tries-before-disconnect {login_times}\r")
    time.sleep(1)
    telnet.execute("end")
    telnet.close   

##################################################
# 차단 시간 확인하기 #

def check_count_and_judge(ip, test_name, fail_log, result_log, judge_count_input):
    j = 0
    count_result = False
    result = "N/A"
    log = []

    telnet = Telnet_Session()
    telnet.connect(ip, username=username_admin2, password=f"{password_admin2}\r", timeout=10)
    telnet.execute("enable\n")
    output = telnet.execute('show syslog\n', timeout=10)
    syslog = output.splitlines()[:30]
    time.sleep(2)
    telnet.execute("q\n")

    for line in syslog:
        if fail_log in line:
            j += 1
            log.append(line)  ##리스트로 만드는것##

        if result_log in line:
            count_result = True
            log.append(line)

    if count_result:
        if j == judge_count_input:
            result = "PASS"
        else:
            result = "FAIL"
        
    print(j)
    final = "\n".join(log)
    print(final)
    
    row1 = get_next_row(sheet1)
    row2 = get_next_row(sheet2)

    sheet1.cell(row=row1, column=2, value=test_name)
    sheet1.cell(row=row1, column=3, value=result)
    sheet1.cell(row=row1, column=4, value=j)

    sheet2.cell(row=row2, column=2, value=test_name)
    sheet2.cell(row=row2, column=3, value=final)

    wb.save(save_path)


##################################################
# 5분 차단 확인하기 #

def check_time_and_judge(ip, test_name, block_log, active_log, judge_time_input):
    j = 0
    count_result = False 
    result = "N/A"
    log_block = ""
    log_active = ""

    telnet = Telnet_Session()
    telnet.connect(ip, username=username_admin2, password=f"{password_admin2}\r", timeout=10)
    telnet.execute("enable\n")
    output = telnet.execute('show syslog\n', timeout=10)
    syslog = output.splitlines()[:30]
    time.sleep(2)
    telnet.execute("q\n")
    
    for line in syslog:
        if block_log in line:
            j += 1
            log_block = line

        if active_log in line:
            count_result = True  ## block_log가 있어야 active_log가 있으므로 block_log가 있는지 확인하는 것
            log_active = line

    time_block = log_block.split()[1]
    time_active = log_active.split()[1]

    dt_block = int(time_block.split(":")[1])
    dt_active = int(time_active.split(":")[1])

    j = dt_active - dt_block

    if count_result:
        if j == judge_time_input:
            result = "PASS"
        else:
            result = "FAIL"
        
    print(j)
    final = "".join(log_block+log_active)
    print(final)
    
    row1 = get_next_row(sheet1)
    row2 = get_next_row(sheet2)

    sheet1.cell(row=row1, column=2, value=test_name)
    sheet1.cell(row=row1, column=3, value=result)
    sheet1.cell(row=row1, column=4, value=j)

    sheet2.cell(row=row2, column=2, value=test_name)
    sheet2.cell(row=row2, column=3, value=final)

    wb.save(save_path)


def connect_ssh(ip, username, wrong_password):
    
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh.connect(hostname=ip, username=username_admin, password=password_admin,
                look_for_keys=False, allow_agent=False)

    shell = ssh.invoke_shell()
    shell.send('enable\n')
    time.sleep(1)

#connect_ssh("172.25.17.73", "admin", "Changeme1357!")

def connect_ssh_retry(ip, username, wrong_password, fail_times, delay=3):
    for i in range(1, fail_times+1):

        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

        try :
            ssh.connect(hostname=ip, username=username, password=wrong_password,
                        look_for_keys=False, allow_agent=False)
        except Exception as e:
            print(f"로그인 실패: {e}")  

        ssh.close()
        time.sleep(delay)

def connect_telnet_retry(ip, username, wrong_password, fail_times, delay=3):
    for i in range(1, fail_times+1):

        telnet = Telnet_Session()
        telnet.enable_debug()

        try : 
            telnet.connect(
                ip, 
                username=username, 
                password=f"{wrong_password}\r", 
                port=23, 
                timeout=5,
                )
        except Exception as e:
            print(f"로그인 실패: {e}")

        telnet.destroy()
        time.sleep(delay)
        
mgmt_ip = "172.25.17.73"
# system_enable(mgmt_ip)
make_user(mgmt_ip, "admin2", 4)
# make_user(mgmt_ip,"user2", 1)
# set_login_time(mgmt_ip, 10)
# set_login_times(mgmt_ip, 3)
# connect_telnet_retry(mgmt_ip, "user2", "wrongpw", 3)
# check_count_and_judge(mgmt_ip, "TEST3_pwfail_count_telnet", "User[user2] failed to Connect.", "%% This account(user2) will be blocked for 10 minutes due to failed logins", 3)
# time.sleep(660)
#check_time_and_judge(mgmt_ip, "TEST4_login_time", "This account(test) will be blocked for 10 minutes", "This account(test) has been activated", 10)


# connect_ssh_retry(mgmt_ip, "user3", "wrongpw", 5)
# check_count_and_judge(mgmt_ip, "TEST3_pwfail_count_ssh", "Failed password for user3 from 10.100.249.100", "%% This account(user3) has been blocked for 5 minutes due to login failure [preauth]", 5)



